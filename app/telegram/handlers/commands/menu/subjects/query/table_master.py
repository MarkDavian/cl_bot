import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from PIL import Image, ImageDraw, ImageFont
from datetime import datetime


class TableMaster:
    def __init__(self, data: list[list]):
        """
        Инициализация мастера таблиц
        
        Args:
            data: Список списков (строк таблицы)
        """
        self.data = data
        self.current_time = datetime.now().strftime("%d.%m.%Y_%H:%M:%S")
        self.xlsx_path = f'files/tmp/query_{self.current_time}.xlsx'
        self.img_path = f'files/tmp/query_{self.current_time}.png'
        
    def _create_xlsx(self) -> None:
        """Создает и стилизует xlsx файл"""
        # Создаем DataFrame и сохраняем в xlsx
        df = pd.DataFrame(self.data)
        df.to_excel(self.xlsx_path, index=False, header=False)
        
        # Открываем файл для стилизации
        wb = openpyxl.load_workbook(self.xlsx_path)
        ws = wb.active
        
        # Настраиваем стили
        header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем стили и авторазмер
        for row in ws.rows:
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row == 1:
                    cell.fill = header_fill
                    
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            
        wb.save(self.xlsx_path)
        
    def _create_image(self) -> None:
        """Создает изображение таблицы"""
        # Создаем изображение таблицы с помощью PIL
        cell_width, cell_height = 320, 100
        table_width = cell_width * len(self.data[0]) 
        table_height = cell_height * len(self.data)

        img = Image.new('RGB', (table_width, table_height), "white")
        draw = ImageDraw.Draw(img)
        font = ImageFont.truetype("/Users/crowly/Library/Fonts/Montserrat-Medium.otf", 28)

        # Отрисовываем ячейки и текст
        for i, row in enumerate(self.data):
            for j, cell in enumerate(row):
                x, y = j * cell_width, i * cell_height
                # Добавляем заливку для заголовков
                if i == 0:
                    draw.rectangle([x, y, x + cell_width, y + cell_height], fill='#E6F3FF', outline="black")
                else:
                    draw.rectangle([x, y, x + cell_width, y + cell_height], outline="black")
                
                # Центрируем текст
                text_size = draw.textbbox((0, 0), str(cell), font=font)
                text_width = text_size[2] - text_size[0]
                text_height = text_size[3] - text_size[1]
                
                draw.text(
                    (x + (cell_width - text_width) / 2, y + (cell_height - text_height) / 2),
                    str(cell),
                    fill="black",
                    font=font
                )

        img.save(self.img_path)
        
        # Обрезаем белые поля
        img = Image.open(self.img_path)
        bg = Image.new(img.mode, img.size, 'white')
        diff = Image.new(img.mode, img.size)
        diff.paste(img)
        bbox = diff.getbbox()
        img_cropped = img.crop(bbox)
        img_cropped.save(self.img_path)
        
    def create_files(self) -> tuple[str, str]:
        """
        Создает xlsx файл и его изображение
        
        Returns:
            tuple[str, str]: Пути к xlsx файлу и его изображению
        """
        self._create_xlsx()
        self._create_image()
        return self.xlsx_path, self.img_path
    
    def delete_files(self) -> None:
        """Удаляет xlsx файл и его изображение"""
        os.remove(self.xlsx_path)
        os.remove(self.img_path)


def create_table_files(data: list[list]) -> tuple[str, str]:
    """
    Функция-обертка для обратной совместимости
    
    Args:
        data: Список списков (строк таблицы)
        
    Returns:
        tuple[str, str]: Пути к xlsx файлу и его изображению
    """
    table_master = TableMaster(data)
    return table_master.create_files()