import shutil
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path

from aiogram import Bot
from aiogram import types, Dispatcher

from config import SETTINGS
from app.core.funcs.get_employees import func_get_employees


async def cmd_export(message: types.Message):
    bot = Bot(SETTINGS.BOT_TOKEN, parse_mode='html')

    employees = func_get_employees()
    
    # Создаем файл Excel с несколькими листами
    excel_file = 'files/Сотрудники.xlsx'
    
    # Создаем writer с engine openpyxl для поддержки нескольких листов
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Основной лист со всеми данными
        emps = [employee.__dict__() for employee in employees]
        df_main = pd.DataFrame.from_dict(emps)
        df_main.to_excel(writer, sheet_name='Все сотрудники', index=False)
        
        # Лист с днями рождения и возрастом (как в birthdays.py)
        data = []
        for emp in employees:
            data.append({
                "ФИО": emp.name,
                "Дата рождения": emp.birthday,
                "Возраст": emp.get_years_old()
            })
        df_birthdays = pd.DataFrame(data)
        df_birthdays.to_excel(writer, sheet_name='Дни рождения', index=False)
        
        # Лист со стажем работы (как в exp_emp.py)
        data = []
        for emp in employees:
            data.append({
                "ФИО": emp.name,
                "Возраст": emp.get_years_old(),
                "Дата начала работы": emp.workstarted,
                "Стаж": emp.get_experience()
            })
        df_exp = pd.DataFrame(data)
        df_exp.to_excel(writer, sheet_name='Стаж работы', index=False)
        
        # Лист с ЛМК (как в expired_lmk.py)
        data = []
        for emp in employees:
            data.append({
                "ФИО": emp.name,
                "Дата ЛМК": emp.lmk,
                "Срок действия": emp.get_lmk_expire_date(),
                "Дней до окончания": emp.time_to_lmk_replacement()
            })
        df_lmk = pd.DataFrame(data)
        df_lmk.to_excel(writer, sheet_name='ЛМК', index=False)
        
        # Лист с сертификатами (как в certificates.py)
        data = []
        for emp in employees:
            data.append({
                "ФИО": emp.name,
                "Базовый": emp.cert_base,
                "Срок базового": emp.get_days_to_cert('base'),
                "Профи": emp.cert_profi,
                "Срок профи": emp.get_days_to_cert('profi')
            })
        df_certs = pd.DataFrame(data)
        df_certs.to_excel(writer, sheet_name='Сертификаты', index=False)
    
    # Стилизуем все листы
    wb = openpyxl.load_workbook(excel_file)
    
    # Настраиваем стили
    header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Применяем стили и авторазмер
        for row in ws.rows:
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row == 1:
                    cell.fill = header_fill
                    
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    
    wb.save(excel_file)
    
    # Создаем zip-архив директории storage
    storage_path = Path('storage')
    zip_path = 'files/storage_backup.zip'
    
    # Удаляем старый архив, если существует
    if os.path.exists(zip_path):
        os.remove(zip_path)
    
    # Создаем новый zip-архив
    shutil.make_archive(
        'files/storage_backup',  # Имя архива без расширения
        'zip',                   # Формат архива
        'storage'                # Директория для архивации
    )
    
    # Отправляем документы в чат
    await bot.send_document(
        chat_id=message.from_user.id,
        document=open(excel_file, "rb"),
        caption="Таблица сотрудников"
    )
    
    # Отправляем zip-архив всей директории storage
    await bot.send_document(
        chat_id=message.from_user.id,
        document=open(zip_path, "rb"),
        caption="Резервная копия директории storage"
    )
    
    # Удаляем файлы после отправки
    os.remove(excel_file)
    os.remove(zip_path)
    
    await bot.session.close()


def reg_export_cmd(dp: Dispatcher):
    dp.register_message_handler(cmd_export, commands="export", state='*')